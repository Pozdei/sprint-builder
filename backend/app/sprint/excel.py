"""Сборка xlsx-выгрузок."""

from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


def _style_header(cell) -> None:
    cell.font = Font(bold=True)
    cell.fill = PatternFill("solid", fgColor="E0E0E0")
    cell.alignment = Alignment(horizontal="center", vertical="center")


def build_sprint_xlsx(
    allocated: list[dict],
    owner_stats: list[dict],
    max_sprint_num: int | None,
    closed_tasks: list[dict | None] | None = None,
    terminal_statuses: list[str] | None = None,
    intrusions: list[dict] | None = None,
) -> bytes:
    """Сборка xlsx спринта.

    Вкладки:
      - Спринт (всегда)
      - Сводка (всегда)
      - Закрытие (только если переданы closed_tasks)
      - Статистика закрытия (только если переданы closed_tasks)
      - Врывы (только если переданы intrusions и не пусто)
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Спринт"

    headers = ["№", "Задача", "Название", "Исполнитель", "Роль", "Бакет",
               "Статус", "Ожид. итог", "Часы", "Sprint", "Приоритет"]
    for col_idx, h in enumerate(headers, start=1):
        _style_header(ws.cell(row=1, column=col_idx, value=h))
    ws.freeze_panes = "A2"

    for row_idx, t in enumerate(allocated, start=2):
        ws.cell(row=row_idx, column=1, value=row_idx - 1)
        if t.get("is_pseudo"):
            cell_key = ws.cell(row=row_idx, column=2, value="(псевдо)")
            cell_key.font = Font(color="888888", italic=True)
        else:
            cell_key = ws.cell(row=row_idx, column=2, value=t.get("key", ""))
            if t.get("url"):
                cell_key.hyperlink = t["url"]
                cell_key.font = Font(color="0563C1", underline="single")

        ws.cell(row=row_idx, column=3, value=t.get("summary", ""))
        ws.cell(row=row_idx, column=4, value=t.get("owner_file_name", ""))
        ws.cell(row=row_idx, column=5, value=t.get("role", ""))
        ws.cell(row=row_idx, column=6, value=t.get("bucket", ""))
        ws.cell(row=row_idx, column=7, value=t.get("status_name", ""))
        ws.cell(row=row_idx, column=8, value=t.get("sprint_expected_result") or "")
        ws.cell(row=row_idx, column=9, value=t.get("hours", 0))
        ws.cell(row=row_idx, column=10, value=t.get("sprint_num", ""))
        ws.cell(row=row_idx, column=11, value=t.get("priority", ""))

    widths = {1: 4, 2: 12, 3: 50, 4: 16, 5: 14, 6: 14, 7: 18, 8: 18, 9: 8, 10: 10, 11: 10}
    for idx, w in widths.items():
        ws.column_dimensions[get_column_letter(idx)].width = w

    # ---- Вкладка "Сводка" ----
    ws_sum = wb.create_sheet("Сводка")
    ws_sum.cell(row=1, column=1, value=f"Спринт {max_sprint_num}" if max_sprint_num else "Спринт").font = Font(bold=True, size=14)

    sum_row = 3
    headers2 = ["Исполнитель", "Часы", "Бюджет", "%"]
    for col_idx, h in enumerate(headers2, start=1):
        _style_header(ws_sum.cell(row=sum_row, column=col_idx, value=h))

    for i, st in enumerate(owner_stats, start=sum_row + 1):
        ws_sum.cell(row=i, column=1, value=st["file_name"])
        ws_sum.cell(row=i, column=2, value=st["used_hours"])
        ws_sum.cell(row=i, column=3, value=st["budget"])
        ws_sum.cell(row=i, column=4, value=f"=B{i}/C{i}")
        ws_sum.cell(row=i, column=4).number_format = "0%"

    for idx, w in {1: 18, 2: 12, 3: 12, 4: 8}.items():
        ws_sum.column_dimensions[get_column_letter(idx)].width = w

    if closed_tasks is not None:
        _add_closure_sheets(wb, allocated, closed_tasks, terminal_statuses or [])

    if intrusions:
        _add_intrusions_sheet(wb, intrusions, terminal_statuses or [])

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _add_closure_sheets(wb, allocated: list[dict],
                         closed_tasks: list[dict | None],
                         terminal_statuses: list[str]) -> None:
    terminal_set = set(terminal_statuses)

    def actual_status(task: dict, closed: dict | None) -> str | None:
        if task.get("is_pseudo"):
            return None
        return closed.get("status_name") if closed else None

    def is_done(task: dict, closed: dict | None) -> bool:
        if task.get("is_pseudo"):
            return True
        st = actual_status(task, closed)
        return st in terminal_set if st else False

    # ---- Вкладка "Закрытие" ----
    ws = wb.create_sheet("Закрытие")
    headers = ["✓?", "Задача", "Название", "Исполнитель", "Роль",
               "Ожид. итог", "Было: статус", "Стало: статус", "Часы"]
    for col_idx, h in enumerate(headers, start=1):
        _style_header(ws.cell(row=1, column=col_idx, value=h))
    ws.freeze_panes = "A2"

    for row_idx, (task, closed) in enumerate(zip(allocated, closed_tasks), start=2):
        done = is_done(task, closed)
        expected = task.get("sprint_expected_result") or ""
        is_pseudo = task.get("is_pseudo", False)

        # Окраска: зелёный если выполнено, жёлтый если ожидали промежуточный этап и достигли, красный иначе
        if done:
            fgColor = "D9EAD3"  # зелёный
        elif expected and not is_pseudo:
            fgColor = "FFF2CC"  # жёлтый — промежуточный результат
        else:
            fgColor = "F4CCCC"  # красный
        fill = PatternFill("solid", fgColor=fgColor)
        for col in range(1, 10):
            ws.cell(row=row_idx, column=col).fill = fill

        ws.cell(row=row_idx, column=1, value="✓" if done else "✗")

        if is_pseudo:
            cell_key = ws.cell(row=row_idx, column=2, value="(псевдо)")
            cell_key.font = Font(color="888888", italic=True)
        else:
            cell_key = ws.cell(row=row_idx, column=2, value=task["key"])
            if task.get("url"):
                cell_key.hyperlink = task["url"]
                cell_key.font = Font(color="0563C1", underline="single")

        ws.cell(row=row_idx, column=3, value=task["summary"])
        ws.cell(row=row_idx, column=4, value=task["owner_file_name"])
        ws.cell(row=row_idx, column=5, value="—" if is_pseudo else task.get("role", ""))
        ws.cell(row=row_idx, column=6, value="—" if is_pseudo else expected)
        ws.cell(row=row_idx, column=7, value="—" if is_pseudo else task["status_name"])
        ws.cell(row=row_idx, column=8,
                value="—" if is_pseudo
                else (closed.get("status_name") if closed else "(нет данных)"))
        ws.cell(row=row_idx, column=9, value=task["hours"])

    widths = {1: 5, 2: 12, 3: 50, 4: 16, 5: 14, 6: 18, 7: 18, 8: 18, 9: 9}
    for idx, w in widths.items():
        ws.column_dimensions[get_column_letter(idx)].width = w

    # ---- Вкладка "Статистика закрытия" ----
    # Считаем по уникальным ключам задач (не строкам), чтобы не двоить pipeline-задачи
    ws_stat = wb.create_sheet("Статистика закрытия")

    # Собираем per-key данные: ключ → (expected_result, actual_status, total_hours)
    key_data: dict[str, dict] = {}
    for t, c in zip(allocated, closed_tasks):
        if t.get("is_pseudo"):
            continue
        k = t["key"]
        if k not in key_data:
            key_data[k] = {
                "summary": t["summary"],
                "expected": t.get("sprint_expected_result") or "",
                "actual": actual_status(t, c),
                "hours": 0.0,
            }
        key_data[k]["hours"] += t["hours"]
        # Обновляем actual при появлении данных
        if key_data[k]["actual"] is None and c:
            key_data[k]["actual"] = c.get("status_name")

    total_keys = len(key_data)
    done_keys = sum(1 for d in key_data.values() if d["actual"] in terminal_set)
    # «По плану» = ожидали терминал и достигли, или ожидали не-терминал и достигли
    expected_done_keys = sum(
        1 for d in key_data.values()
        if d["expected"] in terminal_set and d["actual"] in terminal_set
    )
    exceeded_keys = sum(
        1 for d in key_data.values()
        if d["expected"] not in terminal_set and d["actual"] in terminal_set
    )

    total_hours = sum(t["hours"] for t in allocated)
    done_hours = sum(t["hours"] for t, c in zip(allocated, closed_tasks) if is_done(t, c))

    ws_stat.cell(row=1, column=1, value="Общие метрики (по задачам Jira)").font = Font(bold=True, size=14)

    metrics = [
        ("Уникальных задач", f"{total_keys}", None),
        ("Задач выполнено (терминал)", f"{done_keys} из {total_keys}",
         done_keys / total_keys if total_keys else 0),
        ("Выполнено по плану (ожид. + факт = терминал)", f"{expected_done_keys}",
         expected_done_keys / total_keys if total_keys else 0),
        ("Перевыполнено (ожид. промежуточный, факт терминал)", f"{exceeded_keys}",
         exceeded_keys / total_keys if total_keys else 0),
        ("Часов выполнено (все строки)", f"{round(done_hours, 1)} из {round(total_hours, 1)} ч",
         done_hours / total_hours if total_hours else 0),
    ]
    for i, (label, text, pct) in enumerate(metrics, start=2):
        ws_stat.cell(row=i, column=1, value=label)
        ws_stat.cell(row=i, column=2, value=text)
        if pct is not None:
            ws_stat.cell(row=i, column=3, value=pct).number_format = "0%"

    # Детальная таблица по ключам
    ws_stat.cell(row=8, column=1, value="По задачам (уникальные ключи)").font = Font(bold=True, size=12)
    key_headers = ["Задача", "Ожид. итог", "Факт. статус", "Результат", "Часов всего"]
    for col_idx, h in enumerate(key_headers, start=1):
        _style_header(ws_stat.cell(row=9, column=col_idx, value=h))

    for i, (k, d) in enumerate(key_data.items(), start=10):
        act = d["actual"] or "(нет данных)"
        exp = d["expected"]
        if act in terminal_set:
            verdict = "✓ выполнено"
        elif exp and act and act == exp:
            verdict = "≈ достигли ожидаемого"
        else:
            verdict = "✗ не достигли"
        ws_stat.cell(row=i, column=1, value=k)
        ws_stat.cell(row=i, column=2, value=exp)
        ws_stat.cell(row=i, column=3, value=act)
        ws_stat.cell(row=i, column=4, value=verdict)
        ws_stat.cell(row=i, column=5, value=round(d["hours"], 1))

    for idx, w in {1: 14, 2: 18, 3: 20, 4: 22, 5: 12}.items():
        ws_stat.column_dimensions[get_column_letter(idx)].width = w

    # По людям (часы) — оставляем для детализации нагрузки
    ws_stat.cell(row=max(10 + total_keys + 1, 15), column=1,
                 value="По людям (строки спринта)").font = Font(bold=True, size=12)
    people_row = max(10 + total_keys + 2, 16)

    by_owner: dict[str, dict] = {}
    for t, c in zip(allocated, closed_tasks):
        acc = t["owner_id"]
        if acc not in by_owner:
            by_owner[acc] = {"file_name": t["owner_file_name"],
                              "total": 0, "done": 0, "hours": 0.0, "done_hours": 0.0}
        s = by_owner[acc]
        s["total"] += 1
        s["hours"] += t["hours"]
        if is_done(t, c):
            s["done"] += 1
            s["done_hours"] += t["hours"]

    for col_idx, h in enumerate(["Исполнитель", "Задач", "Часы", "% задач", "% часов"], start=1):
        _style_header(ws_stat.cell(row=people_row, column=col_idx, value=h))

    for i, s in enumerate(by_owner.values(), start=people_row + 1):
        ws_stat.cell(row=i, column=1, value=s["file_name"])
        ws_stat.cell(row=i, column=2, value=f"{s['done']} / {s['total']}")
        ws_stat.cell(row=i, column=3,
                     value=f"{round(s['done_hours'], 1)} / {round(s['hours'], 1)}")
        ws_stat.cell(row=i, column=4,
                     value=s["done"] / s["total"] if s["total"] else 0)
        ws_stat.cell(row=i, column=4).number_format = "0%"
        ws_stat.cell(row=i, column=5,
                     value=s["done_hours"] / s["hours"] if s["hours"] else 0)
        ws_stat.cell(row=i, column=5).number_format = "0%"

    for idx, w in {1: 22, 2: 14, 3: 18, 4: 10, 5: 10}.items():
        ws_stat.column_dimensions[get_column_letter(idx)].width = w


def _add_intrusions_sheet(wb, intrusions: list[dict],
                           terminal_statuses: list[str]) -> None:
    terminal_set = set(terminal_statuses)
    ws = wb.create_sheet("Врывы")

    # Сначала — сводка по (человек, роль)
    ws.cell(row=1, column=1, value="Врывы — сводка").font = Font(bold=True, size=14)

    by_person_role: dict[tuple[str, str], dict] = {}
    for it in intrusions:
        key = (it.get("owner_file_name") or "—", it.get("role") or "")
        if key not in by_person_role:
            by_person_role[key] = {
                "count": 0, "hours": 0.0,
                "done_count": 0, "done_hours": 0.0,
            }
        rec = by_person_role[key]
        rec["count"] += 1
        rec["hours"] += float(it.get("hours") or 0)
        is_done = it.get("is_done") or (it.get("status_name") in terminal_set)
        if is_done:
            rec["done_count"] += 1
            rec["done_hours"] += float(it.get("hours") or 0)

    summary_headers = ["Исполнитель", "Роль", "Задач", "Часы",
                       "Выполнено задач", "Выполнено часов"]
    for col_idx, h in enumerate(summary_headers, start=1):
        _style_header(ws.cell(row=3, column=col_idx, value=h))

    row = 4
    for (file_name, role), rec in sorted(by_person_role.items()):
        ws.cell(row=row, column=1, value=file_name)
        ws.cell(row=row, column=2, value=role)
        ws.cell(row=row, column=3, value=rec["count"])
        ws.cell(row=row, column=4, value=round(rec["hours"], 1))
        ws.cell(row=row, column=5, value=rec["done_count"])
        ws.cell(row=row, column=6, value=round(rec["done_hours"], 1))
        row += 1

    # Подробная таблица
    row += 2
    ws.cell(row=row, column=1, value="Список врывов").font = Font(bold=True, size=12)
    row += 1
    detail_headers = ["✓?", "Задача", "Название", "Исполнитель", "Роль",
                      "Бакет", "Статус", "Часы"]
    for col_idx, h in enumerate(detail_headers, start=1):
        _style_header(ws.cell(row=row, column=col_idx, value=h))
    row += 1

    for it in intrusions:
        status_name = it.get("status_name") or ""
        is_done = it.get("is_done") or (status_name in terminal_set)
        ws.cell(row=row, column=1, value="✓" if is_done else "✗")
        fill = PatternFill("solid", fgColor="D9EAD3" if is_done else "FFF3CD")
        for col in range(1, 9):
            ws.cell(row=row, column=col).fill = fill

        cell_key = ws.cell(row=row, column=2, value=it.get("key", ""))
        if it.get("url"):
            cell_key.hyperlink = it["url"]
            cell_key.font = Font(color="0563C1", underline="single")
        ws.cell(row=row, column=3, value=it.get("summary", ""))
        ws.cell(row=row, column=4, value=it.get("owner_file_name", ""))
        ws.cell(row=row, column=5, value=it.get("role", ""))
        ws.cell(row=row, column=6, value=it.get("bucket", ""))
        ws.cell(row=row, column=7, value=status_name)
        ws.cell(row=row, column=8, value=it.get("hours", 0))
        row += 1

    widths = {1: 5, 2: 12, 3: 50, 4: 18, 5: 14, 6: 14, 7: 18, 8: 9}
    for idx, w in widths.items():
        ws.column_dimensions[get_column_letter(idx)].width = w


_FILL_MISSING_HOURS = PatternFill("solid", fgColor="FFD9D9")   # нежно-красный
_FILL_MISSING_DEV   = PatternFill("solid", fgColor="FFF2CC")   # нежно-жёлтый


_FILL_IN_SPRINT  = PatternFill("solid", fgColor="D9EAD3")   # зелёный — попадёт в спринт


def build_candidates_xlsx(candidates: list[dict],
                           allocated_set: set[str] | None = None) -> bytes:
    """Выгрузка списка кандидатов с оценками, разработчиком и признаком попадания в спринт."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Кандидаты"

    has_alloc = allocated_set is not None
    headers = [
        "Задача", "Название", "Исполнитель", "Роль", "Бакет", "Статус",
        "Часы", "Sprint", "formal_only",
        "Ч. аналитика", "Ч. тестера", "Ч. разработчика", "Разработчик",
    ]
    if has_alloc:
        headers.append("В спринт")

    for col_idx, h in enumerate(headers, start=1):
        _style_header(ws.cell(row=1, column=col_idx, value=h))
    ws.freeze_panes = "A2"

    for row_idx, t in enumerate(candidates, start=2):
        alloc_key = f"{t.get('key')}|{t.get('role')}|{t.get('bucket')}"
        in_sprint = has_alloc and alloc_key in allocated_set

        cell_key = ws.cell(row=row_idx, column=1, value=t.get("key", ""))
        if t.get("url"):
            cell_key.hyperlink = t["url"]
            cell_key.font = Font(color="0563C1", underline="single")
        ws.cell(row=row_idx, column=2, value=t.get("summary", ""))
        ws.cell(row=row_idx, column=3, value=t.get("owner_file_name", ""))
        ws.cell(row=row_idx, column=4, value=t.get("role", ""))
        ws.cell(row=row_idx, column=5, value=t.get("bucket", ""))
        ws.cell(row=row_idx, column=6, value=t.get("status_name", ""))
        ws.cell(row=row_idx, column=7, value=t.get("hours", 0))
        ws.cell(row=row_idx, column=8, value=t.get("sprint_num", ""))
        ws.cell(row=row_idx, column=9, value="да" if t.get("formal_only") else "")

        # Часы по ролям с подсветкой пустых
        for col_offset, field in enumerate(
            ("hours_analyst", "hours_tester", "hours_developer"), start=10
        ):
            val = t.get(field)
            cell = ws.cell(row=row_idx, column=col_offset, value=val or "")
            if not val:
                cell.fill = _FILL_MISSING_HOURS

        # Разработчик с подсветкой отсутствия
        dev_name = t.get("developer_name") or ""
        cell_dev = ws.cell(row=row_idx, column=13, value=dev_name)
        if not dev_name:
            cell_dev.fill = _FILL_MISSING_DEV

        # Признак попадания в спринт
        if has_alloc:
            cell_alloc = ws.cell(row=row_idx, column=14,
                                  value="✓" if in_sprint else "")
            if in_sprint:
                # Подсвечиваем всю строку зелёным
                for col in range(1, 15):
                    ws.cell(row=row_idx, column=col).fill = _FILL_IN_SPRINT
                cell_key.font = Font(color="0563C1", underline="single") if t.get("url") else None

    widths = {
        1: 12, 2: 50, 3: 16, 4: 14, 5: 14, 6: 18,
        7: 8, 8: 10, 9: 12,
        10: 14, 11: 13, 12: 16, 13: 18, 14: 10,
    }
    for idx, w in widths.items():
        ws.column_dimensions[get_column_letter(idx)].width = w

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def build_epic_forecast_xlsx(gantt_items: list[dict], epic_key: str = "") -> bytes:
    """Выгрузка прогноза эпика: 3 листа — по фазам, по задачам, по исполнителям."""
    wb = Workbook()

    # ---- Лист 1: По фазам ----
    ws1 = wb.active
    ws1.title = "По фазам"
    h1 = ["Задача", "Название", "Фаза", "Исполнитель", "Часы", "Стоимость, ₽"]
    for col, h in enumerate(h1, 1):
        _style_header(ws1.cell(row=1, column=col, value=h))
    ws1.freeze_panes = "A2"

    items_sorted = sorted(gantt_items, key=lambda x: (x.get("bucket", ""), x.get("key", "")))
    for row, it in enumerate(items_sorted, 2):
        cell_key = ws1.cell(row=row, column=1, value=it.get("key", ""))
        if it.get("url"):
            cell_key.hyperlink = it["url"]
            cell_key.font = Font(color="0563C1", underline="single")
        ws1.cell(row=row, column=2, value=it.get("summary", ""))
        ws1.cell(row=row, column=3, value=it.get("bucket", ""))
        ws1.cell(row=row, column=4, value=it.get("owner_file_name", ""))
        ws1.cell(row=row, column=5, value=it.get("hours", 0))
        cost = it.get("phase_cost", 0)
        ws1.cell(row=row, column=6, value=cost if cost else "")

    for idx, w in {1: 12, 2: 50, 3: 16, 4: 18, 5: 8, 6: 16}.items():
        ws1.column_dimensions[get_column_letter(idx)].width = w

    # ---- Лист 2: По задачам ----
    ws2 = wb.create_sheet("По задачам")
    h2 = ["Задача", "Название", "Исполнители", "Часов", "Стоимость, ₽"]
    for col, h in enumerate(h2, 1):
        _style_header(ws2.cell(row=1, column=col, value=h))
    ws2.freeze_panes = "A2"

    by_task: dict[str, dict] = {}
    for it in gantt_items:
        k = it.get("key", "")
        if k not in by_task:
            by_task[k] = {
                "summary": it.get("summary", ""),
                "url": it.get("url", ""),
                "executors": set(),
                "hours": 0.0,
                "cost": 0.0,
            }
        by_task[k]["executors"].add(it.get("owner_file_name", ""))
        by_task[k]["hours"] += it.get("hours", 0)
        by_task[k]["cost"] += it.get("phase_cost", 0)

    for row, (k, d) in enumerate(sorted(by_task.items()), 2):
        cell_key = ws2.cell(row=row, column=1, value=k)
        if d["url"]:
            cell_key.hyperlink = d["url"]
            cell_key.font = Font(color="0563C1", underline="single")
        ws2.cell(row=row, column=2, value=d["summary"])
        ws2.cell(row=row, column=3, value=", ".join(sorted(d["executors"])))
        ws2.cell(row=row, column=4, value=round(d["hours"], 1))
        ws2.cell(row=row, column=5, value=round(d["cost"], 0) if d["cost"] else "")

    for idx, w in {1: 12, 2: 50, 3: 35, 4: 10, 5: 16}.items():
        ws2.column_dimensions[get_column_letter(idx)].width = w

    # ---- Лист 3: По исполнителям ----
    ws3 = wb.create_sheet("По исполнителям")
    h3 = ["Исполнитель", "Задач", "Часов", "Стоимость, ₽"]
    for col, h in enumerate(h3, 1):
        _style_header(ws3.cell(row=1, column=col, value=h))
    ws3.freeze_panes = "A2"

    by_person: dict[str, dict] = {}
    for it in gantt_items:
        name = it.get("owner_file_name", "")
        if name not in by_person:
            by_person[name] = {"tasks": set(), "hours": 0.0, "cost": 0.0}
        by_person[name]["tasks"].add(it.get("key", ""))
        by_person[name]["hours"] += it.get("hours", 0)
        by_person[name]["cost"] += it.get("phase_cost", 0)

    for row, (name, d) in enumerate(
        sorted(by_person.items(), key=lambda x: -x[1]["cost"]), 2
    ):
        ws3.cell(row=row, column=1, value=name)
        ws3.cell(row=row, column=2, value=len(d["tasks"]))
        ws3.cell(row=row, column=3, value=round(d["hours"], 1))
        ws3.cell(row=row, column=4, value=round(d["cost"], 0) if d["cost"] else "")

    for idx, w in {1: 22, 2: 10, 3: 10, 4: 16}.items():
        ws3.column_dimensions[get_column_letter(idx)].width = w

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
